"""
AMT Procurement - High Performance SQLite Backend
UPDATED: Smart Logging, Persistent Disk Support, & Cancellation Logic
"""

import os
import json
import hashlib
import shutil
import time
import sqlite3
import io
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl import Workbook
from flask import Flask, jsonify, request, session, send_from_directory, send_file, g
from flask_cors import CORS
from werkzeug.utils import secure_filename

# -------------------------------------------------
# App Init & Configuration
# -------------------------------------------------
app = Flask(__name__, static_folder="static", static_url_path="/static")
app.secret_key = os.environ.get("SECRET_KEY", "AMT_SECRET_KEY_CHANGE_ME_PLEASE")

CORS(app, supports_credentials=True, origins=["*"])

# === PERSISTENT STORAGE CONFIGURATION ===
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Check if running on Render with a persistent disk
if os.path.exists("/var/data"):
    STORAGE_DIR = "/var/data"
else:
    STORAGE_DIR = BASE_DIR

DB_FILE = os.path.join(STORAGE_DIR, "office_ops.db")
BACKUP_DIR = os.path.join(STORAGE_DIR, "backups")
UPLOAD_FOLDER = os.path.join(BASE_DIR, "static", "uploads") # Uploads stay in static for web serving

# Ensure dirs exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)

# Schema Definition
SCHEMAS = {
    "users": """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            password_hash TEXT,
            role TEXT,
            created_at TEXT
        )
    """,
    "requisitions": """
        CREATE TABLE IF NOT EXISTS requisitions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            number TEXT,
            po_number TEXT,
            description TEXT,
            vessel TEXT,
            category TEXT,
            supplier TEXT,
            date_ordered TEXT,
            expected TEXT,
            amount_original REAL,
            currency TEXT,
            amount_usd REAL,
            paid INTEGER,
            delivered INTEGER,
            status TEXT,
            remarks TEXT,
            urgency TEXT,
            tracking_url TEXT,
            created_by TEXT,
            created_at TEXT,
            updated_at TEXT
        )
    """,
    "landings": """
        CREATE TABLE IF NOT EXISTS landings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vessel TEXT,
            item TEXT,
            workshop TEXT,
            expected TEXT,
            landed_date TEXT,
            amount_original REAL,
            currency TEXT,
            amount_usd REAL,
            paid INTEGER,
            delivered INTEGER,
            status TEXT,
            created_by TEXT,
            created_at TEXT,
            updated_at TEXT
        )
    """,
    "directory": """
        CREATE TABLE IF NOT EXISTS directory (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            type TEXT,
            name TEXT,
            email TEXT,
            phone TEXT,
            address TEXT,
            rating INTEGER,
            rating_comment TEXT,
            created_by TEXT,
            created_at TEXT
        )
    """,
    "categories": """
        CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            abbr TEXT,
            created_at TEXT
        )
    """,
    "vessels": """
        CREATE TABLE IF NOT EXISTS vessels (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            created_at TEXT
        )
    """,
    "logs": """
        CREATE TABLE IF NOT EXISTS logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT,
            user TEXT,
            action TEXT,
            target TEXT,
            details TEXT
        )
    """,
    "documents": """
        CREATE TABLE IF NOT EXISTS documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            parent_type TEXT,
            parent_id INTEGER,
            filename TEXT,
            uploaded_at TEXT,
            uploaded_by TEXT
        )
    """
}

DEFAULT_ADMIN = {"username": "admin", "password": "admin123", "role": "admin"}
DEFAULT_FINANCE = {"username": "finance", "password": "finance123", "role": "finance"}

# -------------------------------------------------
# Helpers
# -------------------------------------------------
def get_dubai_time():
    return datetime.utcnow() + timedelta(hours=4)

def now_iso() -> str:
    return get_dubai_time().isoformat(timespec="seconds")

def hash_pw(pw: str) -> str:
    return hashlib.sha256(pw.encode("utf-8")).hexdigest()

# -------------------------------------------------
# Database Logic (SQLite + WAL MODE)
# -------------------------------------------------
def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DB_FILE)
        db.row_factory = sqlite3.Row
        db.execute("PRAGMA journal_mode=WAL")
    return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

def init_db():
    with app.app_context():
        db = get_db()
        for table, schema in SCHEMAS.items():
            db.execute(schema)
        
        # Check Admin
        cur = db.execute("SELECT * FROM users WHERE username = ?", ("admin",))
        if not cur.fetchone():
            db.execute("INSERT INTO users (username, password_hash, role, created_at) VALUES (?, ?, ?, ?)",
                       ("admin", hash_pw(DEFAULT_ADMIN["password"]), "admin", now_iso()))
        
        # Check Finance
        cur = db.execute("SELECT * FROM users WHERE username = ?", ("finance",))
        if not cur.fetchone():
            db.execute("INSERT INTO users (username, password_hash, role, created_at) VALUES (?, ?, ?, ?)",
                       ("finance", hash_pw(DEFAULT_FINANCE["password"]), "finance", now_iso()))
        
        db.commit()

def query_db(query, args=(), one=False):
    cur = get_db().execute(query, args)
    rv = cur.fetchall()
    cur.close()
    return (rv[0] if rv else None) if one else rv

def modify_db(query, args=()):
    db = get_db()
    cur = db.execute(query, args)
    db.commit()
    last_id = cur.lastrowid
    cur.close()
    return last_id

# -------------------------------------------------
# Auth Helpers
# -------------------------------------------------
def current_user() -> Optional[Dict[str, str]]:
    if "username" not in session: return None
    return {"username": session["username"], "role": session.get("role", "user")}

def log_action(action: str, target: str = "", details: str = "") -> None:
    try:
        u = current_user()
        username = u["username"] if u else "system"
        modify_db("INSERT INTO logs (timestamp, user, action, target, details) VALUES (?, ?, ?, ?, ?)",
                  (now_iso(), username, action, str(target), str(details)))
    except Exception as e:
        print(f"Logging error: {e}")

def require_login():
    if not current_user(): return jsonify({"error": "login_required"}), 401

def require_admin():
    u = current_user()
    if not u: return jsonify({"error": "login_required"}), 401
    if u["role"] != "admin": return jsonify({"error": "admin_required"}), 403

def require_write():
    u = current_user()
    if not u: return jsonify({"error": "login_required"}), 401
    if u["role"] == "finance": return jsonify({"error": "read_only"}), 403

# -------------------------------------------------
# FX / Utils
# -------------------------------------------------
def fetch_fx_rates():
    return {"USD": 1.0, "EUR": 0.95, "AED": 3.673, "GBP": 0.79, "SGD": 1.35}

def to_usd(amount, currency):
    try: val = float(amount)
    except: return 0.0
    currency = (currency or "USD").upper()
    if currency == "USD": return val
    rates = fetch_fx_rates()
    rate = rates.get(currency, 1.0)
    return val / rate if rate else val

def save_db_to_excel(filepath):
    wb = Workbook()
    if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
    
    with app.app_context():
        db = get_db()
        for tbl in SCHEMAS.keys():
            ws = wb.create_sheet(tbl)
            cur = db.execute(f"SELECT * FROM {tbl}")
            rows = cur.fetchall()
            if cur.description:
                headers = [d[0] for d in cur.description]
                ws.append(headers)
                for r in rows: ws.append(list(r))
            cur.close()
    wb.save(filepath)

# -------------------------------------------------
# Routes
# -------------------------------------------------
@app.before_request
def check_init():
    if not getattr(app, '_init_done', False):
        init_db()
        app._init_done = True

@app.route("/")
def home():
    return send_from_directory("static", "index.html")

@app.route("/api/health")
def health():
    return jsonify({"status": "ok", "time": now_iso()})

# --- Auth ---
@app.post("/api/login")
def login():
    data = request.json or {}
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    
    user = query_db("SELECT * FROM users WHERE username = ?", (username,), one=True)
    
    if not user or user["password_hash"] != hash_pw(password):
        return jsonify({"error": "invalid_credentials"}), 401
        
    session["username"] = username
    session["role"] = user["role"]
    log_action("Login", details="User logged in")
    return jsonify({"ok": True, "username": username, "role": session["role"]})

@app.post("/api/logout")
def logout():
    session.clear()
    return jsonify({"ok": True})

@app.get("/api/session")
def get_session():
    u = current_user()
    if not u: return jsonify({"logged_in": False}), 401
    return jsonify({"logged_in": True, **u})

@app.get("/api/currencies")
def get_currencies():
    return jsonify({"currencies": sorted(list(fetch_fx_rates().keys()))})

# --- Requisitions ---
@app.get("/api/requisitions")
def list_reqs():
    rows = query_db("SELECT * FROM requisitions")
    return jsonify([dict(r) for r in rows])

@app.post("/api/requisitions")
def add_req():
    if require_write(): return require_write()
    d = request.json or {}
    
    amt = d.get("amount_original")
    if amt in [None, ""]: amt = d.get("amount", 0)
    curr = d.get("currency", "USD")
    amt_usd = round(to_usd(amt, curr), 2)
    
    po = d.get("po_number")
    ref = d.get("number")
    
    sql = """
        INSERT INTO requisitions (number, po_number, description, vessel, category, supplier,
        date_ordered, expected, amount_original, currency, amount_usd, paid, delivered, status, 
        remarks, urgency, tracking_url, created_by, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'open', ?, ?, ?, ?, ?)
    """
    
    rid = modify_db(sql, (
        ref, po, d.get("description"), d.get("vessel"),
        d.get("category"), d.get("supplier"), d.get("date_ordered"), d.get("expected"),
        amt, curr, amt_usd, 1 if d.get("paid") else 0, int(d.get("delivered", 0)),
        d.get("remarks"), d.get("urgency"), d.get("tracking_url"),
        current_user()["username"], now_iso()
    ))
    
    new_row = query_db("SELECT * FROM requisitions WHERE id = ?", (rid,), one=True)
    log_action("Create Order", target=po or ref or str(rid), details=f"Created new order for {d.get('vessel')}")
    return jsonify(dict(new_row))

@app.patch("/api/requisitions/<int:rid>")
def edit_req(rid):
    if require_write(): return require_write()
    d = request.json or {}
    
    # --- SMART LOGGING LOGIC ---
    action_log = "Update Order"
    details_log = "Updated details"
    
    # Check specifically for status changes (Cancellation/Recovery)
    if "status" in d:
        if d["status"] == "cancelled":
            action_log = "Cancel Order"
            details_log = "Order moved to bin"
        elif d["status"] == "open":
            action_log = "Recover Order"
            details_log = "Order recovered from bin"
            
    # Check specifically for marking paid via single edit
    elif "paid" in d:
        action_log = "Update Payment"
        details_log = "Marked as Paid" if d["paid"] else "Marked as Unpaid"
    # ---------------------------

    fields = []
    values = []
    
    if "amount_original" in d or "amount" in d or "currency" in d:
        amt = d.get("amount_original") or d.get("amount") or 0
        curr = d.get("currency", "USD")
        d["amount_usd"] = round(to_usd(amt, curr), 2)
        d["amount_original"] = amt
    
    if "paid" in d: d["paid"] = 1 if d["paid"] else 0
    
    allowed_cols = ["po_number", "number", "description", "vessel", "category", "supplier",
                    "date_ordered", "expected", "amount_original", "currency", "amount_usd",
                    "paid", "delivered", "status", "remarks", "urgency", "tracking_url"]
                    
    for k, v in d.items():
        if k in allowed_cols:
            fields.append(f"{k} = ?")
            values.append(v)
            
    if not fields: return jsonify({"ok": True})
    
    fields.append("updated_at = ?")
    values.append(now_iso())
    values.append(rid)
    
    modify_db(f"UPDATE requisitions SET {', '.join(fields)} WHERE id = ?", tuple(values))
    
    # Get PO number for better logging
    row = query_db("SELECT po_number, number FROM requisitions WHERE id=?", (rid,), one=True)
    target_ref = row['po_number'] or row['number'] or str(rid) if row else str(rid)
    
    log_action(action_log, target=target_ref, details=details_log)
    return jsonify({"ok": True})

@app.delete("/api/requisitions/<int:rid>")
def del_req(rid):
    if require_admin(): return require_admin()
    
    # Get info before delete for log
    row = query_db("SELECT po_number, number FROM requisitions WHERE id=?", (rid,), one=True)
    target_ref = row['po_number'] or row['number'] or str(rid) if row else str(rid)
    
    modify_db("DELETE FROM requisitions WHERE id = ?", (rid,))
    log_action("Delete Order", target=target_ref, details="Permanently deleted from system")
    return jsonify({"ok": True})

@app.post("/api/requisitions/bulk")
def bulk_req():
    if require_write(): return require_write()
    d = request.json or {}
    ids = d.get("ids", [])
    action = d.get("action")
    
    if not ids: return jsonify({"ok": False})
    
    sql = ""
    log_action_name = "Bulk Action"
    
    if action == "mark_paid": 
        sql = "UPDATE requisitions SET paid = 1 WHERE id = ?"
        log_action_name = "Bulk Pay"
    elif action == "mark_unpaid": 
        sql = "UPDATE requisitions SET paid = 0 WHERE id = ?"
        log_action_name = "Bulk Unpay"
    elif action == "mark_delivered": 
        sql = "UPDATE requisitions SET delivered = 1 WHERE id = ?"
        log_action_name = "Bulk Receive"
    elif action == "mark_partial": 
        sql = "UPDATE requisitions SET delivered = 2 WHERE id = ?"
        log_action_name = "Bulk Partial"
    else: return jsonify({"error": "invalid"}), 400
    
    db = get_db()
    for i in ids:
        db.execute(sql, (i,))
    db.commit()
    
    log_action(log_action_name, target=f"{len(ids)} Orders", details=f"Applied {action} to {len(ids)} items")
        
    return jsonify({"ok": True})

# --- Landings ---
@app.get("/api/landings")
def list_landings():
    return jsonify([dict(r) for r in query_db("SELECT * FROM landings")])

@app.post("/api/landings")
def add_landing():
    if require_write(): return require_write()
    d = request.json or {}
    
    amt = d.get("amount_original") or d.get("amount") or 0
    curr = d.get("currency", "USD")
    amt_usd = round(to_usd(amt, curr), 2)
    
    sql = """INSERT INTO landings (vessel, item, workshop, expected, landed_date, 
             amount_original, currency, amount_usd, paid, delivered, status, created_by, created_at)
             VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'open', ?, ?)"""
             
    lid = modify_db(sql, (d.get("vessel"), d.get("item"), d.get("workshop"), d.get("expected"),
                          d.get("landed_date"), amt, curr, amt_usd, 1 if d.get("paid") else 0,
                          int(d.get("delivered",0)), current_user()["username"], now_iso()))
                          
    log_action("Add Landing", target=d.get("item"), details=f"Added landing for {d.get('vessel')}")
    return jsonify({"ok": True})

@app.patch("/api/landings/<int:lid>")
def edit_landing(lid):
    if require_write(): return require_write()
    d = request.json or {}
    
    fields, values = [], []
    if "amount_original" in d or "amount" in d:
        amt = d.get("amount_original") or d.get("amount") or 0
        curr = d.get("currency", "USD")
        d["amount_usd"] = round(to_usd(amt, curr), 2)
        d["amount_original"] = amt
        
    if "paid" in d: d["paid"] = 1 if d["paid"] else 0
    
    allowed = ["vessel", "item", "workshop", "expected", "landed_date", "amount_original",
               "currency", "amount_usd", "paid", "delivered"]
               
    for k, v in d.items():
        if k in allowed:
            fields.append(f"{k} = ?")
            values.append(v)
            
    if fields:
        fields.append("updated_at = ?")
        values.append(now_iso())
        values.append(lid)
        modify_db(f"UPDATE landings SET {', '.join(fields)} WHERE id = ?", tuple(values))
    
    log_action("Edit Landing", target=str(lid))
    return jsonify({"ok": True})

@app.delete("/api/landings/<int:lid>")
def del_landing(lid):
    if require_admin(): return require_admin()
    modify_db("DELETE FROM landings WHERE id = ?", (lid,))
    log_action("Delete Landing", target=str(lid))
    return jsonify({"ok": True})

@app.post("/api/landings/bulk")
def bulk_land():
    if require_write(): return require_write()
    d = request.json or {}
    ids = d.get("ids", [])
    action = d.get("action")
    
    sql = ""
    if action == "mark_paid": sql = "UPDATE landings SET paid = 1 WHERE id = ?"
    elif action == "mark_unpaid": sql = "UPDATE landings SET paid = 0 WHERE id = ?"
    elif action == "mark_delivered": sql = "UPDATE landings SET delivered = 1 WHERE id = ?"
    elif action == "mark_partial": sql = "UPDATE landings SET delivered = 2 WHERE id = ?"
    else: return jsonify({"error": "invalid"}), 400
    
    db = get_db()
    for i in ids: db.execute(sql, (i,))
    db.commit()
    
    log_action("Bulk Landing Action", target=f"{len(ids)} Items", details=action)
    return jsonify({"ok": True})

# --- Directory, Categories, Vessels ---
@app.get("/api/directory")
def list_dir():
    return jsonify([dict(r) for r in query_db("SELECT * FROM directory ORDER BY type")])

@app.post("/api/directory")
def add_dir():
    if require_write(): return require_write()
    d = request.json
    modify_db("INSERT INTO directory (type, name, email, phone, address, rating, created_by, created_at) VALUES (?,?,?,?,?,?,?,?)",
              (d.get("type"), d.get("name"), d.get("email"), d.get("phone"), d.get("address"), d.get("rating",5),
               current_user()["username"], now_iso()))
    log_action("Add Contact", target=d.get("name"), details=d.get("type"))
    return jsonify({"ok": True})

@app.patch("/api/directory/<int:did>")
def edit_dir(did):
    if require_write(): return require_write()
    d = request.json
    fields, vals = [], []
    for k in ["name", "email", "phone", "address", "rating", "rating_comment"]:
        if k in d:
            fields.append(f"{k}=?")
            vals.append(d[k])
    if fields:
        vals.append(did)
        modify_db(f"UPDATE directory SET {','.join(fields)} WHERE id=?", tuple(vals))
    log_action("Edit Contact", target=str(did))
    return jsonify({"ok": True})

@app.delete("/api/directory/<int:did>")
def del_dir(did):
    if require_admin(): return require_admin()
    modify_db("DELETE FROM directory WHERE id=?", (did,))
    log_action("Delete Contact", target=str(did))
    return jsonify({"ok": True})

@app.get("/api/categories")
def get_cats(): return jsonify([dict(r) for r in query_db("SELECT * FROM categories")])

@app.post("/api/categories")
def add_cat():
    if require_admin(): return require_admin()
    modify_db("INSERT INTO categories (name, abbr, created_at) VALUES (?,?,?)",
              (request.json.get("name"), request.json.get("abbr"), now_iso()))
    log_action("Add Category", target=request.json.get("name"))
    return jsonify({"ok": True})

@app.delete("/api/categories/<int:cid>")
def del_cat(cid):
    if require_admin(): return require_admin()
    modify_db("DELETE FROM categories WHERE id=?", (cid,))
    return jsonify({"ok": True})

@app.get("/api/vessels")
def get_ves(): return jsonify([dict(r) for r in query_db("SELECT * FROM vessels")])

@app.post("/api/vessels")
def add_ves():
    if require_admin(): return require_admin()
    modify_db("INSERT INTO vessels (name, created_at) VALUES (?,?)",
              (request.json.get("name"), now_iso()))
    log_action("Add Vessel", target=request.json.get("name"))
    return jsonify({"ok": True})

@app.delete("/api/vessels/<int:vid>")
def del_ves(vid):
    if require_admin(): return require_admin()
    modify_db("DELETE FROM vessels WHERE id=?", (vid,))
    return jsonify({"ok": True})

@app.get("/api/users")
def get_users():
    if require_admin(): return require_admin()
    return jsonify([dict(r) for r in query_db("SELECT id, username, role, created_at FROM users")])

@app.post("/api/users")
def add_user():
    if require_admin(): return require_admin()
    d = request.json
    try:
        modify_db("INSERT INTO users (username, password_hash, role, created_at) VALUES (?,?,?,?)",
                  (d.get("username"), hash_pw(d.get("password")), d.get("role","user"), now_iso()))
        log_action("Create User", target=d.get("username"), details=f"Role: {d.get('role')}")
        return jsonify({"ok": True})
    except sqlite3.IntegrityError:
        return jsonify({"error": "duplicate_user"}), 409

@app.delete("/api/users/<username>")
def del_user(username):
    if require_admin(): return require_admin()
    if username == "admin": return jsonify({"error": "cannot_delete_root"}), 400
    modify_db("DELETE FROM users WHERE username=?", (username,))
    log_action("Delete User", target=username)
    return jsonify({"ok": True})

# --- Logs & Reports ---
@app.get("/api/audit")
def get_logs():
    if require_admin(): return require_admin()
    rows = query_db("SELECT user, action, target, details, timestamp as date FROM logs ORDER BY id DESC LIMIT 500")
    return jsonify([dict(r) for r in rows])

@app.get("/api/reports/aging")
def aging_report():
    if require_login(): return require_login()
    rows = query_db("SELECT * FROM requisitions WHERE paid=0 AND status != 'cancelled'")
    now = get_dubai_time()
    out = []
    for r in rows:
        try: dt = datetime.strptime(str(r["date_ordered"]).split("T")[0], "%Y-%m-%d")
        except: continue
        
        delta = (now - dt).days
        grp = "< 30 Days"
        if delta > 90: grp = "> 90 Days"
        elif delta > 60: grp = "60-90 Days"
        elif delta > 30: grp = "30-60 Days"
        
        out.append({
            "po": r["po_number"] or r["number"],
            "supplier": r["supplier"],
            "amount": r["amount_usd"],
            "days": delta,
            "group": grp
        })
    return jsonify(sorted(out, key=lambda x: x["days"], reverse=True))

# --- Documents ---
@app.post("/api/documents/upload")
def upload_doc():
    if require_write(): return require_write()
    f = request.files.get("file")
    if not f: return jsonify({"error": "no_file"}), 400
    
    fname = secure_filename(f.filename)
    save_name = f"{int(time.time())}_{fname}"
    f.save(os.path.join(UPLOAD_FOLDER, save_name))
    
    modify_db("INSERT INTO documents (parent_type, parent_id, filename, uploaded_at, uploaded_by) VALUES (?,?,?,?,?)",
              (request.form.get("parent_type"), request.form.get("parent_id"), save_name, now_iso(), current_user()["username"]))
    log_action("Upload Doc", target=save_name)
    return jsonify({"ok": True})

@app.get("/api/documents/<ptype>/<pid>")
def get_docs(ptype, pid):
    if require_login(): return require_login()
    rows = query_db("SELECT * FROM documents WHERE parent_type=? AND parent_id=?", (ptype, pid))
    return jsonify([dict(r) for r in rows])

# -------------------------------------------------
# EXCEL IMPORT / EXPORT
# -------------------------------------------------

@app.get("/api/backup/download")
def download_excel_backup():
    if require_admin(): return require_admin()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"backup_{ts}.xlsx"
    fpath = os.path.join(BACKUP_DIR, fname)
    save_db_to_excel(fpath)
    return send_file(fpath, as_attachment=True, download_name=fname)

@app.post("/api/backup/create")
def create_backup_internal():
    if require_admin(): return require_admin()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"backup_{ts}.xlsx"
    fpath = os.path.join(BACKUP_DIR, fname)
    save_db_to_excel(fpath)
    log_action("Create Backup", target=fname)
    return jsonify({"ok": True})

@app.get("/api/backups")
def list_backups():
    if require_admin(): return require_admin()
    out = []
    if os.path.exists(BACKUP_DIR):
        for f in sorted(os.listdir(BACKUP_DIR), reverse=True):
            if f.endswith(".xlsx"):
                p = os.path.join(BACKUP_DIR, f)
                st = os.stat(p)
                out.append({
                    "name": f,
                    "size": st.st_size,
                    "created_at": datetime.fromtimestamp(st.st_mtime).isoformat()
                })
    return jsonify(out)

@app.get("/api/backups/<name>/download")
def download_specific_backup(name):
    if require_admin(): return require_admin()
    safe_name = secure_filename(name)
    return send_from_directory(BACKUP_DIR, safe_name, as_attachment=True)

@app.post("/api/backups/<name>/restore")
def restore_backup_file(name):
    if require_admin(): return require_admin()
    safe_name = secure_filename(name)
    src = os.path.join(BACKUP_DIR, safe_name)
    if not os.path.exists(src): return jsonify({"error": "not_found"}), 404
    
    with open(src, "rb") as f:
        try:
            wb = openpyxl.load_workbook(src)
        except:
            return jsonify({"error": "invalid_excel"}), 400
            
        db = get_db()
        for tbl in SCHEMAS.keys():
            if tbl in wb.sheetnames:
                ws = wb[tbl]
                rows = list(ws.iter_rows(values_only=True))
                if not rows: continue
                headers = [h for h in rows[0] if h]
                data = rows[1:]
                if not data: continue
                
                db.execute(f"DELETE FROM {tbl}")
                placeholders = ",".join(["?"] * len(headers))
                cols = ",".join(headers)
                sql = f"INSERT OR IGNORE INTO {tbl} ({cols}) VALUES ({placeholders})"
                
                for r in data:
                    try: db.execute(sql, r[:len(headers)])
                    except: pass
        db.commit()
    
    log_action("Restore Backup", target=safe_name, details="Overwrote DB with backup")
    return jsonify({"ok": True})

@app.post("/api/upload")
def restore_from_excel():
    if require_admin(): return require_admin()
    if "file" not in request.files: return jsonify({"error": "no_file"}), 400
    
    f = request.files["file"]
    try: wb = openpyxl.load_workbook(f)
    except: return jsonify({"error": "invalid_excel"}), 400
        
    db = get_db()
    for tbl in SCHEMAS.keys():
        if tbl in wb.sheetnames:
            ws = wb[tbl]
            rows = list(ws.iter_rows(values_only=True))
            if not rows: continue
            
            headers = [h for h in rows[0] if h]
            data = rows[1:]
            
            if not data: continue
            
            db.execute(f"DELETE FROM {tbl}")
            placeholders = ",".join(["?"] * len(headers))
            cols = ",".join(headers)
            sql = f"INSERT OR IGNORE INTO {tbl} ({cols}) VALUES ({placeholders})"
            
            for r in data:
                try: db.execute(sql, r[:len(headers)])
                except: pass
                    
    db.commit()
    log_action("Restore DB", details="Restored from Excel upload")
    return jsonify({"ok": True})

if __name__ == "__main__":
    init_db()
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
